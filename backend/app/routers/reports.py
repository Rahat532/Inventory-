from fastapi import APIRouter, Depends, HTTPException, Response
from sqlalchemy.orm import Session
from sqlalchemy import func, desc
from datetime import datetime
from typing import Optional
import io
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import csv

from app.database import get_db
from app.models import Product, Sale, Category
from app.schemas import ReportRequest

router = APIRouter()
def _safe_sheet_title(title: str) -> str:
    # Excel constraints: max 31 chars, cannot contain: : \\ / ? * [ ]
    forbidden = set(':\\/?*[]')
    # Prefer a short base (before parentheses)
    base = (title.split('(')[0] or title).strip()
    cleaned = ''.join(ch for ch in base if ch not in forbidden)
    cleaned = cleaned.strip()
    if not cleaned:
        cleaned = 'Sheet'
    if len(cleaned) > 31:
        cleaned = cleaned[:31]
    return cleaned



def create_pdf_report(title: str, headers: list, data: list, filename: str = None) -> bytes:
    """Create a PDF report with given data"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    
    # Title style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    # Create content
    content = []
    
    # Add title
    content.append(Paragraph(title, title_style))
    content.append(Spacer(1, 20))
    
    # Add generation date
    content.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    content.append(Spacer(1, 20))
    
    # Create table
    if data:
        # Prepare table data with headers
        table_data = [headers] + data
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        content.append(table)
    else:
        content.append(Paragraph("No data available for the selected criteria.", styles['Normal']))
    
    # Build PDF
    doc.build(content)
    buffer.seek(0)
    return buffer.getvalue()


def create_excel_report(title: str, headers: list, data: list, filename: str = None) -> bytes:
    """Create an Excel report with given data"""
    buffer = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = _safe_sheet_title(title)
    
    # Set up styles
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add title
    last_col = get_column_letter(len(headers))
    worksheet.merge_cells(f'A1:{last_col}1')
    worksheet['A1'] = title
    worksheet['A1'].font = Font(bold=True, size=16)
    worksheet['A1'].alignment = Alignment(horizontal="center")
    
    # Add generation date
    worksheet.merge_cells(f'A2:{last_col}2')
    worksheet['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    worksheet['A2'].alignment = Alignment(horizontal="center")
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        # Set background color
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Add data
    for row_idx, row_data in enumerate(data, 5):
        for col_idx, value in enumerate(row_data, 1):
            worksheet.cell(row=row_idx, column=col_idx, value=value)

    # Apply number formats for known numeric columns
    header_to_currency = {"Price", "Cost", "Total Amount", "Discount", "Tax", "Final Amount"}
    header_to_integer = {"Stock Qty", "Min Stock", "Deficit", "Product Count", "Total Stock"}
    currency_cols = [i + 1 for i, h in enumerate(headers) if h in header_to_currency]
    integer_cols = [i + 1 for i, h in enumerate(headers) if h in header_to_integer]
    last_row = worksheet.max_row
    for col_idx in currency_cols:
        for row_idx in range(5, last_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            # Only set number format; assume creator passed numeric values for currency columns
            cell.number_format = '[$$-409]#,##0.00'
            cell.alignment = Alignment(horizontal="right")
    for col_idx in integer_cols:
        for row_idx in range(5, last_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.number_format = '0'
    
    # Auto-adjust column widths (avoid merged cell issue)
    header_row = 4
    last_row = worksheet.max_row
    num_cols = len(headers)
    for col_idx in range(1, num_cols + 1):
        max_length = 0
        for row_idx in range(header_row, last_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            try:
                value_len = len(str(cell.value)) if cell.value is not None else 0
                if value_len > max_length:
                    max_length = value_len
            except Exception:
                continue
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def create_csv_report(headers: list, data: list) -> bytes:
    """Create a CSV report (UTF-8 with BOM for Excel compatibility)"""
    output = io.StringIO()
    writer = csv.writer(output, lineterminator='\r\n')
    writer.writerow(headers)
    for row in data:
        writer.writerow(row)
    text = output.getvalue()
    # Prepend UTF-8 BOM for better Excel handling of UTF-8
    return ('\ufeff' + text).encode('utf-8')


@router.post("/generate")
async def generate_report(request: ReportRequest, db: Session = Depends(get_db)):
    """Generate a report based on the request"""
    
    # Parse dates if provided
    start_date = None
    end_date = None
    
    def _parse_date(date_str: Optional[str], end: bool = False) -> Optional[datetime]:
        if not date_str:
            return None
        try:
            dt = datetime.fromisoformat(date_str)
        except Exception:
            return None
        # If date string without time (YYYY-MM-DD), normalize to start or end of day
        if len(date_str) <= 10:
            if end:
                dt = dt.replace(hour=23, minute=59, second=59, microsecond=999999)
            else:
                dt = dt.replace(hour=0, minute=0, second=0, microsecond=0)
        return dt

    start_date = _parse_date(request.start_date, end=False)
    end_date = _parse_date(request.end_date, end=True)
    
    # Generate report based on type
    if request.report_type == "sales":
        return await generate_sales_report(start_date, end_date, request.format, db)
    elif request.report_type == "inventory":
        return await generate_inventory_report(request.format, db)
    elif request.report_type == "categories":
        return await generate_categories_report(request.format, db)
    elif request.report_type == "low_stock":
        return await generate_low_stock_report(request.format, db)
    else:
        raise HTTPException(status_code=400, detail="Invalid report type")


async def generate_sales_report(start_date: Optional[datetime], end_date: Optional[datetime], format: str, db: Session):
    """Generate sales report"""
    
    # Build query
    query = db.query(
        Sale.sale_number,
        Sale.created_at,
        Sale.total_amount,
        Sale.discount,
        Sale.tax,
        Sale.final_amount,
        Sale.payment_method
    )
    
    if start_date:
        query = query.filter(Sale.created_at >= start_date)
    if end_date:
        query = query.filter(Sale.created_at <= end_date)
    
    sales = query.order_by(desc(Sale.created_at)).all()
    
    # Prepare data
    headers = ["Sale Number", "Date", "Total Amount", "Discount", "Tax", "Final Amount", "Payment Method"]
    if format == "pdf":
        data = []
        for sale in sales:
            data.append([
                sale.sale_number,
                sale.created_at.strftime("%Y-%m-%d %H:%M"),
                f"${sale.total_amount:.2f}",
                f"${sale.discount:.2f}",
                f"${sale.tax:.2f}",
                f"${sale.final_amount:.2f}",
                sale.payment_method
            ])
    else:
        data = []
        for sale in sales:
            data.append([
                sale.sale_number,
                sale.created_at.strftime("%Y-%m-%d %H:%M"),
                round(float(sale.total_amount or 0), 2),
                round(float(sale.discount or 0), 2),
                round(float(sale.tax or 0), 2),
                round(float(sale.final_amount or 0), 2),
                sale.payment_method
            ])
    
    # Generate title
    title = "Sales Report"
    if start_date and end_date:
        title += f" ({start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')})"
    elif start_date:
        title += f" (From {start_date.strftime('%Y-%m-%d')})"
    elif end_date:
        title += f" (Until {end_date.strftime('%Y-%m-%d')})"
    
    # Generate report
    if format == "pdf":
        pdf_data = create_pdf_report(title, headers, data)
        return Response(
            content=pdf_data,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"}
        )
    elif format == "excel":  # Excel
        excel_data = create_excel_report(title, headers, data)
        return Response(
            content=excel_data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
        )
    else:  # CSV
        csv_data = create_csv_report(headers, data)
        return Response(
            content=csv_data,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
        )


async def generate_inventory_report(format: str, db: Session):
    """Generate inventory report"""
    
    # Get all active products with category info
    products = db.query(
        Product.name,
        Product.sku,
        Category.name.label('category_name'),
        Product.stock_quantity,
        Product.min_stock_level,
        Product.price,
        Product.cost,
        Product.unit
    ).join(Category, Product.category_id == Category.id)\
     .filter(Product.is_active)\
     .order_by(Category.name, Product.name)\
     .all()
    
    # Prepare data
    headers = ["Product Name", "SKU", "Category", "Stock Qty", "Min Stock", "Price", "Cost", "Unit"]
    if format == "pdf":
        data = []
        for product in products:
            data.append([
                product.name,
                product.sku,
                product.category_name,
                product.stock_quantity,
                product.min_stock_level,
                f"${product.price:.2f}",
                f"${product.cost:.2f}",
                product.unit
            ])
    else:
        data = []
        for product in products:
            data.append([
                product.name,
                product.sku,
                product.category_name,
                int(product.stock_quantity or 0),
                int(product.min_stock_level or 0),
                round(float(product.price or 0), 2),
                round(float(product.cost or 0), 2),
                product.unit
            ])
    
    title = "Inventory Report"
    
    # Generate report
    if format == "pdf":
        pdf_data = create_pdf_report(title, headers, data)
        return Response(
            content=pdf_data,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=inventory_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"}
        )
    elif format == "excel":  # Excel
        excel_data = create_excel_report(title, headers, data)
        return Response(
            content=excel_data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=inventory_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
        )
    else:  # CSV
        csv_data = create_csv_report(headers, data)
        return Response(
            content=csv_data,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=inventory_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
        )


async def generate_categories_report(format: str, db: Session):
    """Generate categories report"""
    
    # Get categories with product counts
    categories = db.query(
        Category.name,
        Category.description,
        func.count(Product.id).label('product_count'),
        func.sum(Product.stock_quantity).label('total_stock'),
        func.avg(Product.price).label('avg_price')
    ).outerjoin(Product, Category.id == Product.category_id)\
     .filter(Product.is_active.is_(True) | Product.is_active.is_(None))\
     .group_by(Category.id, Category.name, Category.description)\
     .order_by(Category.name)\
     .all()
    
    # Prepare data
    headers = ["Category", "Description", "Product Count", "Total Stock", "Avg Price"]
    data = []
    
    for category in categories:
        data.append([
            category.name,
            category.description or "N/A",
            category.product_count or 0,
            category.total_stock or 0,
            f"${category.avg_price:.2f}" if category.avg_price else "$0.00"
        ])
    
    title = "Categories Report"
    
    # Generate report
    if format == "pdf":
        pdf_data = create_pdf_report(title, headers, data)
        return Response(
            content=pdf_data,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=categories_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"}
        )
    elif format == "excel":  # Excel
        excel_data = create_excel_report(title, headers, data)
        return Response(
            content=excel_data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=categories_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
        )
    else:  # CSV
        csv_data = create_csv_report(headers, data)
        return Response(
            content=csv_data,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=categories_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
        )


async def generate_low_stock_report(format: str, db: Session):
    """Generate low stock report (products below minimum stock)"""
    products = db.query(
        Product.name,
        Product.sku,
        Category.name.label('category_name'),
        Product.stock_quantity,
        Product.min_stock_level,
        Product.price,
        Product.unit
    ).join(Category, Product.category_id == Category.id)\
     .filter(Product.is_active.is_(True))\
     .filter(Product.stock_quantity < Product.min_stock_level)\
     .order_by(Category.name, Product.name)\
     .all()

    headers = ["Product Name", "SKU", "Category", "Stock Qty", "Min Stock", "Deficit", "Price", "Unit"]
    if format == "pdf":
        data = []
        for p in products:
            deficit = (p.min_stock_level or 0) - (p.stock_quantity or 0)
            data.append([
                p.name,
                p.sku,
                p.category_name,
                p.stock_quantity,
                p.min_stock_level,
                max(deficit, 0),
                f"${p.price:.2f}",
                p.unit
            ])
    else:
        data = []
        for p in products:
            deficit = (p.min_stock_level or 0) - (p.stock_quantity or 0)
            data.append([
                p.name,
                p.sku,
                p.category_name,
                int(p.stock_quantity or 0),
                int(p.min_stock_level or 0),
                max(int(deficit), 0),
                round(float(p.price or 0), 2),
                p.unit
            ])

    title = "Low Stock Report"
    if format == "pdf":
        pdf_data = create_pdf_report(title, headers, data)
        return Response(
            content=pdf_data,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=low_stock_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"}
        )
    elif format == "excel":
        excel_data = create_excel_report(title, headers, data)
        return Response(
            content=excel_data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=low_stock_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
        )
    else:
        csv_data = create_csv_report(headers, data)
        return Response(
            content=csv_data,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=low_stock_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
        )